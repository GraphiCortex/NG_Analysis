#!/usr/bin/env python3
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ─── 1) Define paths ──────────────────────────────────────────────────────────
root       = Path(__file__).parents[1]               # …/Project
processed  = root / "Data" / "Processed"
expr_int   = processed / "expression_integration"
limma_dir  = processed / "limma_results"
out_xlsx   = expr_int / "E2F4_GO_overview.xlsx"

# Sheet 1 source
master_fp  = Path("Project/Data/Processed/expression_integration/integration") / "binding_expression_master.csv"

# ORA symbol files for the 312 terms
ora_cell_fp = Path("Project/Data/Processed/limma_results") / "E2F4_Cell_cycle_ORA_symbols.csv"
ora_dna_fp  = Path("Project/Data/Processed/limma_results") / "E2F4_DNA_repair_ORA_symbols.csv"
ora_apo_fp  = Path("Project/Data/Processed/limma_results") / "E2F4_Cell_death_ORA_symbols.csv"
# ───────────────────────────────────────────────────────────────────────────────

# ─── 2) Sheet 1: the binding–expression master (all rows) ───────────────────────
master = pd.read_csv(master_fp)
# Keep exactly the columns in your CSV:
sheet1 = master[[
    "Gene",    # gene symbol
    "LFC",     # log₂FC from expression
    "adjP",    # adjusted p-value
    "sig",     # TRUE/FALSE significance flag
    "category",# expression GO category
    "M_rep1", "M_rep2", "M_rep3", "M_avg"
]].rename(columns={
    "Gene":       "gene_symbol",
    "LFC":        "log2FC",
    "adjP":       "adj.P.Val",
    "sig":        "significant"
})
# ───────────────────────────────────────────────────────────────────────────────

# ─── 3) Sheet 2: all 312 GO terms, with category column ────────────────────────
def load_ora(fp, cat_label):
    df = pd.read_csv(fp)[["Description"]].copy()
    df["Category"] = cat_label
    return df

ora_cell = load_ora(ora_cell_fp, "Cell-cycle")
ora_dna  = load_ora(ora_dna_fp,  "DNA-repair")
ora_apo  = load_ora(ora_apo_fp,  "Apoptosis")

sheet2 = pd.concat([ora_cell, ora_dna, ora_apo], ignore_index=True).drop_duplicates()

# ─── 4) Sheets 3–5: term → gene breakdown by category ──────────────────────────
def explode_terms(fp, cat_label):
    df = pd.read_csv(fp)[["Description","geneID"]]
    df = df.assign(geneID=df["geneID"].str.split("/")).explode("geneID")
    df["geneID"] = df["geneID"].str.strip()
    df["Category"] = cat_label
    return df.rename(columns={"geneID":"gene_symbol"})

tg_cell = explode_terms(ora_cell_fp, "Cell-cycle")
tg_dna  = explode_terms(ora_dna_fp,  "DNA-repair")
tg_apo  = explode_terms(ora_apo_fp,  "Apoptosis")

sheet3 = tg_cell[tg_cell["Category"]=="Cell-cycle"][["Description","gene_symbol"]]
sheet4 = tg_dna [tg_dna ["Category"]=="DNA-repair"][["Description","gene_symbol"]]
sheet5 = tg_apo [tg_apo ["Category"]=="Apoptosis"][["Description","gene_symbol"]]

# ─── 5) Write a 5‐sheet Excel workbook ─────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

def add_sheet(name, df, highlight_col=None, color_map=None):
    ws = wb.create_sheet(name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    if highlight_col and color_map:
        idx = df.columns.get_loc(highlight_col) + 1
        for cell in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            val = cell[0].value
            if val in color_map:
                cell[0].fill = PatternFill(fill_type="solid", start_color=color_map[val])

# Sheet 1
add_sheet("1045_genes", sheet1)

# Sheet 2 with color‐coded categories
fills = {
    "Cell-cycle": "FFFF99",
    "DNA-repair": "CCFFCC",
    "Apoptosis":  "FFCCCC",
}
add_sheet("312_GO_terms", sheet2, highlight_col="Category", color_map=fills)

# Sheets 3–5
add_sheet("Cell_cycle_terms",    sheet3)
add_sheet("DNA_repair_terms",    sheet4)
add_sheet("Apoptosis_terms",     sheet5)

wb.save(out_xlsx)
print("✅ Written workbook to:", out_xlsx)
