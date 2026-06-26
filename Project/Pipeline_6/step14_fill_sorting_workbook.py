#!/usr/bin/env python3
import os
import glob
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# Paths
# =========================

# Existing workbook to edit
matches = glob.glob(os.path.join("Project", "sorting apoptotic genes fong oshikawa*.xlsx"))
matches = [m for m in matches if not os.path.basename(m).startswith("~$")]

if not matches:
    raise FileNotFoundError("Could not find 'sorting apoptotic genes fong oshikawa*.xlsx' under Project/")

WORKBOOK_PATH = matches[0]

# Fong ranked files
FONG_FILES = {
    "Apoptosis": [
        "Project/New_Data/new_data/apoptosis_candidates_ranked(t).csv",
        "Project/New_Data/apoptosis_candidates_ranked.csv",
    ],
    "Autophagy": [
        "Project/New_Data/new_data/autophagy_candidates_ranked(t).csv",
        "Project/New_Data/autophagy_candidates_ranked.csv",
    ],
    "Necroptosis": [
        "Project/New_Data/new_data/necroptosis_candidates_ranked(t).csv",
        "Project/New_Data/necroptosis_candidates_ranked.csv",
    ],
}

# CAG ranked files
CAG_FILES = {
    "Apoptosis": "Project/New_Data_5/new_data/apoptosis_candidates_ranked_cag.csv",
    "Autophagy": "Project/New_Data_5/new_data/autophagy_candidates_ranked_cag.csv",
    "Necroptosis": "Project/New_Data_5/new_data/necroptosis_candidates_ranked_cag.csv",
}

# MAP2 ranked files
MAP2_FILES = {
    "Apoptosis": "Project/New_Data_4/new_data/apoptosis_candidates_ranked_map2.csv",
    "Autophagy": "Project/New_Data_4/new_data/autophagy_candidates_ranked_map2.csv",
    "Necroptosis": "Project/New_Data_4/new_data/necroptosis_candidates_ranked_map2.csv",
}

# Andrusiak apoptosis ranked file
ANDR_APOPTOSIS = "Project/New_Data_6/new_data/apoptosis_candidates_ranked_andrusiak.csv"


# =========================
# Helpers
# =========================

def pick_existing(paths):
    if isinstance(paths, str):
        paths = [paths]
    for p in paths:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(f"None of these paths exist: {paths}")


def load_ranked(path):
    df = pd.read_csv(path)
    if "Gene" not in df.columns:
        raise ValueError(f"{path} is missing a Gene column.")

    # Expression column can be named in our usual style
    expr_col = None
    for c in ["mean_logFC_expr", "log2FoldChange", "logFC_expr"]:
        if c in df.columns:
            expr_col = c
            break

    if expr_col is None:
        raise ValueError(f"{path} has no recognizable expression logFC column.")

    out = df[["Gene", expr_col]].copy()
    out["Gene"] = out["Gene"].astype(str).str.upper().str.strip()
    out[expr_col] = pd.to_numeric(out[expr_col], errors="coerce")

    # If duplicate genes exist, keep strongest expression signal
    out["abs_expr_tmp"] = out[expr_col].abs()
    idx = out.groupby("Gene")["abs_expr_tmp"].idxmax()
    out = out.loc[idx, ["Gene", expr_col]].rename(columns={expr_col: "expr"})
    return out.set_index("Gene")["expr"].to_dict()


def build_table(family, include_andrusiak=False, base_from_sheet=None):
    fong = load_ranked(pick_existing(FONG_FILES[family]))
    cag  = load_ranked(pick_existing(CAG_FILES[family]))
    map2 = load_ranked(pick_existing(MAP2_FILES[family]))

    andr = {}
    if include_andrusiak:
        andr = load_ranked(pick_existing(ANDR_APOPTOSIS))

    # For the Andrusiak apoptosis sheet, use the first apoptosis sheet as base gene order if available
    base_genes = []
    if base_from_sheet is not None:
        for row in base_from_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            g = row[0]
            if g is not None and str(g).strip():
                base_genes.append(str(g).upper().strip())

    all_genes = set(fong) | set(cag) | set(map2) | set(andr)

    # Preserve base order first, then add remaining genes by strongest available expression
    ordered = []
    seen = set()

    for g in base_genes:
        if g in all_genes and g not in seen:
            ordered.append(g)
            seen.add(g)

    remaining = list(all_genes - seen)

    def max_abs_expr(g):
        vals = []
        for d in [fong, cag, map2, andr]:
            if g in d and pd.notna(d[g]):
                vals.append(abs(d[g]))
        return max(vals) if vals else 0

    remaining = sorted(remaining, key=max_abs_expr, reverse=True)
    genes = ordered + remaining

    if include_andrusiak:
        rows = []
        for g in genes:
            rows.append([
                g,
                1 if g in fong else 0,
                1 if g in cag else 0,
                1 if g in map2 else 0,
                1 if g in andr else 0,
                fong.get(g, None),
                cag.get(g, None),
                map2.get(g, None),
                andr.get(g, None),
            ])
        headers = [
            "Gene", "fong", "cag", "map2", "andrusiak",
            "expr_fong", "expr_cag", "expr_map2", "expr_andrusiak"
        ]
    else:
        rows = []
        for g in genes:
            rows.append([
                g,
                1 if g in fong else 0,
                1 if g in cag else 0,
                1 if g in map2 else 0,
                fong.get(g, None),
                cag.get(g, None),
                map2.get(g, None),
            ])
        headers = [
            "Gene", "fong", "cag", "map2",
            "expr_fong", "expr_cag", "expr_map2"
        ]

    return headers, rows


def find_sheet(wb, contains_terms):
    """
    Find sheet whose name contains all terms, case-insensitive.
    """
    terms = [t.lower() for t in contains_terms]
    for ws in wb.worksheets:
        name = ws.title.lower()
        if all(t in name for t in terms):
            return ws
    raise ValueError(f"Could not find sheet containing terms: {contains_terms}. Existing sheets: {wb.sheetnames}")


def write_table(ws, headers, rows):
    # Clear only the generated table area A:K, leave any role/function columns further right alone
    max_clear_col = max(len(headers), 12)
    for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 2000), min_col=1, max_col=max_clear_col):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border()

    # Write headers
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j).value = h

    # Write rows
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j).value = val

    style_sheet(ws, len(headers), len(rows) + 1)


def style_sheet(ws, ncols, nrows):
    # Colors
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    presence_fill = PatternFill("solid", fgColor="C6EFCE")
    missing_expr_fill = PatternFill("solid", fgColor="A6A6A6")
    negative_fill = PatternFill("solid", fgColor="F4A6A6")
    positive_fill = PatternFill("solid", fgColor="C6EFCE")
    thin = Side(style="thin", color="D9D9D9")
    medium = Side(style="medium", color="000000")

    # Header
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=medium)

    # Column groups
    # presence columns = B:D, or B:E if Andrusiak included
    has_andr = "andrusiak" in [ws.cell(1, c).value for c in range(1, ncols + 1)]
    presence_start = 2
    presence_end = 5 if has_andr else 4
    expr_start = presence_end + 1
    expr_end = ncols

    for r in range(2, nrows + 1):
        # gene name
        ws.cell(r, 1).alignment = Alignment(horizontal="left")

        # presence cells
        for c in range(presence_start, presence_end + 1):
            cell = ws.cell(r, c)
            if cell.value == 1:
                cell.fill = presence_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # expression cells
        for c in range(expr_start, expr_end + 1):
            cell = ws.cell(r, c)
            if cell.value is None or cell.value == "":
                cell.fill = missing_expr_fill
            else:
                try:
                    v = float(cell.value)
                    cell.number_format = "0.000000"
                    if v < 0:
                        cell.fill = negative_fill
                    elif v > 0:
                        # light green only for positives; remove this block if you want white positives
                        cell.fill = positive_fill
                except Exception:
                    pass

        # light grid borders
        for c in range(1, ncols + 1):
            ws.cell(r, c).border = Border(bottom=thin)

    # Thick separators like the template
    sep_cols = [1, presence_end, expr_end]
    for c in sep_cols:
        for r in range(1, nrows + 1):
            ws.cell(r, c).border = Border(
                right=medium,
                bottom=ws.cell(r, c).border.bottom
            )

    # Widths
    widths = {
        1: 26,
        2: 14,
        3: 14,
        4: 14,
        5: 14 if has_andr else 18,
    }

    for c in range(1, ncols + 1):
        if c in widths:
            ws.column_dimensions[get_column_letter(c)].width = widths[c]
        elif c >= expr_start:
            ws.column_dimensions[get_column_letter(c)].width = 16
        else:
            ws.column_dimensions[get_column_letter(c)].width = 14

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"


# =========================
# Main
# =========================

print(f"Using workbook: {WORKBOOK_PATH}")

# Backup first
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
backup_path = WORKBOOK_PATH.replace(".xlsx", f"_backup_{timestamp}.xlsx")
shutil.copy2(WORKBOOK_PATH, backup_path)
print(f"Backup created: {backup_path}")

wb = load_workbook(WORKBOOK_PATH)

# Find sheets
ws_apop = find_sheet(wb, ["apoptosis"])
ws_necro = find_sheet(wb, ["necroptosis"])
ws_auto = find_sheet(wb, ["autophagy"])
ws_andr = find_sheet(wb, ["andruisak"]) if any("andruisak" in s.lower() for s in wb.sheetnames) else find_sheet(wb, ["andrusiak"])

# Fill necroptosis
headers, rows = build_table("Necroptosis", include_andrusiak=False)
write_table(ws_necro, headers, rows)
print(f"Filled sheet: {ws_necro.title} ({len(rows)} rows)")

# Fill autophagy
headers, rows = build_table("Autophagy", include_andrusiak=False)
write_table(ws_auto, headers, rows)
print(f"Filled sheet: {ws_auto.title} ({len(rows)} rows)")

# Fill Andrusiak apoptosis, using first apoptosis sheet order as base
headers, rows = build_table("Apoptosis", include_andrusiak=True, base_from_sheet=ws_apop)
write_table(ws_andr, headers, rows)
print(f"Filled sheet: {ws_andr.title} ({len(rows)} rows)")

# Save in place
wb.save(WORKBOOK_PATH)
print(f"✓ Updated workbook in place: {WORKBOOK_PATH}")